from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta, date
from accounts.decorators import roles_requis
from produits.models import Produit, StockAnnexe
from stocks.models import MouvementStock
from facturation.models import Facture
from caisse.models import Paiement
from livraisons.models import Tournee
from django.db.models import Sum
import calendar


def get_donnees_rapport(date_debut, date_fin):
    mouvements = MouvementStock.objects.filter(
        date__date__range=[date_debut, date_fin]
    ).select_related('produit', 'annexe', 'effectue_par')

    paiements = Paiement.objects.filter(
        date__date__range=[date_debut, date_fin]
    )

    factures = Facture.objects.filter(
        date_creation__date__range=[date_debut, date_fin]
    ).select_related('client', 'annexe')

    total_encaisse = sum(p.montant for p in paiements)
    total_cash = sum(p.montant for p in paiements if p.mode_paiement == 'cash')
    total_mtn = sum(p.montant for p in paiements if p.mode_paiement == 'mtn_momo')
    total_orange = sum(p.montant for p in paiements if p.mode_paiement == 'orange_money')

    entrees = mouvements.filter(type_mouvement='entree')
    sorties = mouvements.filter(type_mouvement__in=['sortie', 'vente_directe'])
    retours = mouvements.filter(type_mouvement='retour')

    stocks_bas = StockAnnexe.objects.filter(
        quantite__lte=10
    ).select_related('produit', 'annexe')

    factures_impayees = Facture.objects.filter(
        statut='en_attente'
    ).select_related('client')

    return {
        'entrees': entrees,
        'sorties': sorties,
        'retours': retours,
        'factures': factures,
        'paiements': paiements,
        'total_encaisse': total_encaisse,
        'total_cash': total_cash,
        'total_mtn': total_mtn,
        'total_orange': total_orange,
        'stocks_bas': stocks_bas,
        'factures_impayees': factures_impayees,
    }


@login_required
@roles_requis('responsable', 'superviseur')
def rapport_journalier(request):
    aujourd_hui = timezone.now().date()
    date_str = request.GET.get('date', str(aujourd_hui))
    try:
        date_rapport = date.fromisoformat(date_str)
    except ValueError:
        date_rapport = aujourd_hui

    donnees = get_donnees_rapport(date_rapport, date_rapport)
    tournees = Tournee.objects.filter(
        date=date_rapport
    ).select_related('livreur', 'annexe')

    return render(request, 'rapports/journalier.html', {
        'date_rapport': date_rapport,
        'aujourd_hui': aujourd_hui,
        'tournees': tournees,
        **donnees,
    })


@login_required
@roles_requis('responsable', 'superviseur')
def rapport_hebdomadaire(request):
    aujourd_hui = timezone.now().date()
    debut_semaine = aujourd_hui - timedelta(days=aujourd_hui.weekday())
    fin_semaine = debut_semaine + timedelta(days=6)

    ventes_semaine = []
    for i in range(7):
        jour = debut_semaine + timedelta(days=i)
        paiements_jour = Paiement.objects.filter(date__date=jour)
        total = sum(p.montant for p in paiements_jour)
        ventes_semaine.append({
            'jour': jour,
            'total': total,
            'nb_factures': Facture.objects.filter(date_creation__date=jour).count(),
        })

    top_produits = MouvementStock.objects.filter(
        date__date__range=[debut_semaine, fin_semaine],
        type_mouvement__in=['sortie', 'vente_directe']
    ).values('produit__nom').annotate(
        total_sorti=Sum('quantite')
    ).order_by('-total_sorti')[:5]

    paiements_semaine = Paiement.objects.filter(
        date__date__range=[debut_semaine, fin_semaine]
    )
    total_semaine = sum(p.montant for p in paiements_semaine)

    return render(request, 'rapports/hebdomadaire.html', {
        'debut_semaine': debut_semaine,
        'fin_semaine': fin_semaine,
        'ventes_semaine': ventes_semaine,
        'top_produits': top_produits,
        'total_semaine': total_semaine,
        'aujourd_hui': aujourd_hui,
    })


@login_required
@roles_requis('responsable', 'superviseur')
def rapport_mensuel(request):
    aujourd_hui = timezone.now().date()
    mois = int(request.GET.get('mois', aujourd_hui.month))
    annee = int(request.GET.get('annee', aujourd_hui.year))

    premier_jour = date(annee, mois, 1)
    dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])

    # Ventes par semaine du mois
    donnees = get_donnees_rapport(premier_jour, dernier_jour)

    # Ventes par jour du mois
    ventes_par_jour = []
    for jour_num in range(1, dernier_jour.day + 1):
        jour = date(annee, mois, jour_num)
        paiements_jour = Paiement.objects.filter(date__date=jour)
        total = sum(p.montant for p in paiements_jour)
        ventes_par_jour.append({
            'jour': jour,
            'total': total,
            'nb_factures': Facture.objects.filter(
                date_creation__date=jour
            ).count(),
        })

    # Top produits du mois
    top_produits = MouvementStock.objects.filter(
        date__date__range=[premier_jour, dernier_jour],
        type_mouvement__in=['sortie', 'vente_directe']
    ).values('produit__nom').annotate(
        total_sorti=Sum('quantite')
    ).order_by('-total_sorti')[:10]

    # Liste des mois et années pour le sélecteur
    mois_liste = [
        (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
        (5, 'Mai'), (6, 'Juin'), (7, 'Juillet'), (8, 'Août'),
        (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre')
    ]
    annees_liste = list(range(aujourd_hui.year - 2, aujourd_hui.year + 1))

    return render(request, 'rapports/mensuel.html', {
        'premier_jour': premier_jour,
        'dernier_jour': dernier_jour,
        'mois': mois,
        'annee': annee,
        'mois_liste': mois_liste,
        'annees_liste': annees_liste,
        'ventes_par_jour': ventes_par_jour,
        'top_produits': top_produits,
        'aujourd_hui': aujourd_hui,
        **donnees,
    })


# ============================================================
# EXPORTS
# ============================================================

@login_required
@roles_requis('responsable', 'superviseur')
def export_pdf(request, type_rapport):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer

    aujourd_hui = timezone.now().date()

    if type_rapport == 'journalier':
        date_str = request.GET.get('date', str(aujourd_hui))
        try:
            date_rapport = date.fromisoformat(date_str)
        except ValueError:
            date_rapport = aujourd_hui
        titre = f"Rapport Journalier — {date_rapport.strftime('%d/%m/%Y')}"
        donnees = get_donnees_rapport(date_rapport, date_rapport)

    elif type_rapport == 'mensuel':
        mois = int(request.GET.get('mois', aujourd_hui.month))
        annee = int(request.GET.get('annee', aujourd_hui.year))
        premier_jour = date(annee, mois, 1)
        dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])
        titre = f"Rapport Mensuel — {calendar.month_name[mois]} {annee}"
        donnees = get_donnees_rapport(premier_jour, dernier_jour)
    else:
        titre = "Rapport"
        donnees = get_donnees_rapport(aujourd_hui, aujourd_hui)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="rapport_{type_rapport}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4,
                            topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    BLEU = colors.HexColor('#1A376C')
    VERT = colors.HexColor('#1E7E34')
    ROUGE = colors.HexColor('#C05514')

    elements = []

    # Titre
    titre_style = ParagraphStyle(
        'Titre', parent=styles['Title'],
        textColor=BLEU, fontSize=18, spaceAfter=6
    )
    elements.append(Paragraph("DepôtGest", titre_style))
    elements.append(Paragraph(titre, styles['Heading2']))
    elements.append(Paragraph(
        f"Généré le {aujourd_hui.strftime('%d/%m/%Y')} par {request.user.get_full_name()}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5*cm))

    # Résumé financier
    elements.append(Paragraph("Résumé Financier", styles['Heading3']))
    data_finance = [
        ['Mode de paiement', 'Montant (FCFA)'],
        ['Espèces', f"{donnees['total_cash']:,.0f}"],
        ['MTN Mobile Money', f"{donnees['total_mtn']:,.0f}"],
        ['Orange Money', f"{donnees['total_orange']:,.0f}"],
        ['TOTAL ENCAISSÉ', f"{donnees['total_encaisse']:,.0f}"],
    ]
    table_finance = Table(data_finance, colWidths=[10*cm, 7*cm])
    table_finance.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BLEU),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D6E4F0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(table_finance)
    elements.append(Spacer(1, 0.5*cm))

    # Mouvements de stock
    elements.append(Paragraph("Mouvements de Stock", styles['Heading3']))
    data_mvt = [['Produit', 'Type', 'Annexe', 'Quantité', 'Heure']]
    for m in donnees['entrees']:
        data_mvt.append([
            m.produit.nom, 'Entrée', str(m.annexe),
            f"+{m.quantite}", m.date.strftime('%H:%M')
        ])
    for m in donnees['sorties']:
        data_mvt.append([
            m.produit.nom, 'Sortie', str(m.annexe),
            f"-{m.quantite}", m.date.strftime('%H:%M')
        ])
    for m in donnees['retours']:
        data_mvt.append([
            m.produit.nom, 'Retour', str(m.annexe),
            f"+{m.quantite}", m.date.strftime('%H:%M')
        ])

    if len(data_mvt) > 1:
        table_mvt = Table(data_mvt, colWidths=[5*cm, 3*cm, 3*cm, 3*cm, 3*cm])
        table_mvt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BLEU),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table_mvt)
    else:
        elements.append(Paragraph("Aucun mouvement.", styles['Normal']))

    elements.append(Spacer(1, 0.5*cm))

    # Factures impayées
    if donnees['factures_impayees']:
        elements.append(Paragraph("Factures Impayées", styles['Heading3']))
        data_imp = [['Numéro', 'Client', 'Montant restant (FCFA)']]
        for f in donnees['factures_impayees']:
            data_imp.append([f.numero, str(f.client), f"{f.montant_restant:,.0f}"])
        table_imp = Table(data_imp, colWidths=[6*cm, 7*cm, 5*cm])
        table_imp.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ROUGE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5F5')]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table_imp)

    doc.build(elements)
    return response


@login_required
@roles_requis('responsable', 'superviseur')
def export_excel(request, type_rapport):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    aujourd_hui = timezone.now().date()

    if type_rapport == 'journalier':
        date_str = request.GET.get('date', str(aujourd_hui))
        try:
            date_rapport = date.fromisoformat(date_str)
        except ValueError:
            date_rapport = aujourd_hui
        titre = f"Rapport Journalier {date_rapport.strftime('%d-%m-%Y')}"
        donnees = get_donnees_rapport(date_rapport, date_rapport)

    elif type_rapport == 'mensuel':
        mois = int(request.GET.get('mois', aujourd_hui.month))
        annee = int(request.GET.get('annee', aujourd_hui.year))
        premier_jour = date(annee, mois, 1)
        dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])
        titre = f"Rapport Mensuel {calendar.month_name[mois]} {annee}"
        donnees = get_donnees_rapport(premier_jour, dernier_jour)
    else:
        titre = "Rapport"
        donnees = get_donnees_rapport(aujourd_hui, aujourd_hui)

    wb = openpyxl.Workbook()

    # Couleurs
    BLEU = "1A376C"
    VERT = "1E7E34"
    ORANGE = "C05514"
    BLEU_CLAIR = "D6E4F0"

    def style_header(cell, bg=BLEU):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def border_cell(cell):
        thin = Side(style='thin', color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Feuille 1 : Résumé financier ----
    ws1 = wb.active
    ws1.title = "Résumé Financier"

    ws1.merge_cells('A1:C1')
    ws1['A1'] = f"DepôtGest — {titre}"
    ws1['A1'].font = Font(bold=True, size=14, color=BLEU)
    ws1['A2'] = f"Généré le {aujourd_hui.strftime('%d/%m/%Y')}"
    ws1['A2'].font = Font(italic=True, color="888888")
    ws1.append([])

    headers = ['Mode de paiement', 'Montant (FCFA)']
    ws1.append(headers)
    for cell in ws1[ws1.max_row]:
        style_header(cell)

    rows_finance = [
        ('Espèces', float(donnees['total_cash'])),
        ('MTN Mobile Money', float(donnees['total_mtn'])),
        ('Orange Money', float(donnees['total_orange'])),
        ('TOTAL ENCAISSÉ', float(donnees['total_encaisse'])),
    ]
    for i, row in enumerate(rows_finance):
        ws1.append(list(row))
        for cell in ws1[ws1.max_row]:
            border_cell(cell)
            if i == len(rows_finance) - 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill("solid", fgColor=BLEU_CLAIR)

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20

    # ---- Feuille 2 : Mouvements ----
    ws2 = wb.create_sheet("Mouvements de Stock")
    headers2 = ['Produit', 'Type', 'Annexe', 'Quantité', 'Effectué par', 'Heure']
    ws2.append(headers2)
    for cell in ws2[1]:
        style_header(cell)

    for m in donnees['entrees']:
        ws2.append([
            m.produit.nom, 'Entrée fournisseur', str(m.annexe),
            m.quantite, m.effectue_par.get_full_name() if m.effectue_par else '—',
            m.date.strftime('%H:%M')
        ])
        ws2[ws2.max_row][3].font = Font(color="1E7E34", bold=True)

    for m in donnees['sorties']:
        ws2.append([
            m.produit.nom, 'Sortie', str(m.annexe),
            m.quantite, m.effectue_par.get_full_name() if m.effectue_par else '—',
            m.date.strftime('%H:%M')
        ])
        ws2[ws2.max_row][3].font = Font(color="C05514", bold=True)

    for m in donnees['retours']:
        ws2.append([
            m.produit.nom, 'Retour', str(m.annexe),
            m.quantite, m.effectue_par.get_full_name() if m.effectue_par else '—',
            m.date.strftime('%H:%M')
        ])

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[col].width = 18

    # ---- Feuille 3 : Factures impayées ----
    ws3 = wb.create_sheet("Factures Impayées")
    headers3 = ['Numéro', 'Client', 'Montant Total', 'Payé', 'Reste à payer', 'Date']
    ws3.append(headers3)
    for cell in ws3[1]:
        style_header(cell, bg=ORANGE)

    for f in donnees['factures_impayees']:
        ws3.append([
            f.numero, str(f.client),
            float(f.montant_total), float(f.montant_paye),
            float(f.montant_restant),
            f.date_creation.strftime('%d/%m/%Y')
        ])
        ws3[ws3.max_row][4].font = Font(color="C01414", bold=True)

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 20

    # ---- Feuille 4 : Stocks bas ----
    ws4 = wb.create_sheet("Stocks Bas")
    headers4 = ['Produit', 'Annexe', 'Quantité actuelle', 'Seuil alerte']
    ws4.append(headers4)
    for cell in ws4[1]:
        style_header(cell, bg="5a189a")

    for s in donnees['stocks_bas']:
        ws4.append([str(s.produit), str(s.annexe), s.quantite, s.seuil_alerte])
        ws4[ws4.max_row][2].font = Font(color="C01414", bold=True)

    for col in ['A', 'B', 'C', 'D']:
        ws4.column_dimensions[col].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapport_{type_rapport}.xlsx"'
    wb.save(response)
    return response